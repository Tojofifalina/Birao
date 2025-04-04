from django.http import  HttpResponse,HttpResponseRedirect
from django.shortcuts import render, get_object_or_404,redirect
from django.contrib.auth.decorators import login_required,permission_required
from datetime import datetime,date
from  django.core.paginator import Paginator
import openpyxl
from django.shortcuts import render
from django.db.models import Q,Sum,Avg
from datetime import date
from django.db.models import Sum, Avg, Q,Value,IntegerField
from django.contrib import messages
from Fitatanana.models import Mpino3,Adidy,Katekomen,Fanamarihana
from Vola.models import VolaNiditra,VolaNivoka
import json
import re
from django.db.models.functions import Cast
from django.http import JsonResponse
from .form import FanamarihanaForm

def fitre_special(laharana):
    isa = re.findall(r'\d+',laharana)
    return [int(p) for p in isa] if isa else [float('inf')]

@login_required
def Mpino(request):
    moba = Mpino3.objects.all()

    fikaroana = request.GET.get('q', '')
    if not fikaroana:
        fikaroana = date.year
    isa = Mpino3.objects.all().count()
    ita = Mpino3.objects.filter(Anarana__icontains=fikaroana) if fikaroana else None


    # Récupérer tous les paramètres de filtre et de recherche
    fikaroana = request.GET.get('q', '')
    taona_nahatogavana = request.GET.get('taona', '')
    nanoboka = request.GET.get('nanoboka', '')
    nifarana = request.GET.get('nifarana', '')
    par_page = int(request.GET.get('par_page', 10))

    Mpadray = request.GET.get('mpadray', '')

    # Filtrer les personnes
    #mpino = Mpino3.objects.all().order_by("LaharanaKaratra")
    #mpino = sorted(Mpino3.objects.all(),key=lambda m: fitre_special(m.LaharanaKaratra))
    mpino = Mpino3.objects.all().annotate(num_laharana=Cast('LaharanaKaratra',IntegerField())).order_by('num_laharana')
    if fikaroana:
        mpino = mpino.filter(Anarana__icontains=fikaroana) or mpino.filter(LaharanaKaratra__icontains=str(fikaroana)) if fikaroana else None
    if taona_nahatogavana:
        mpino = mpino.filter(TaonaNahatogavanaTaoAmpiagonana=taona_nahatogavana)
    if nanoboka and nifarana:
        mpino = mpino.filter(TaonaNahatogavanaTaoAmpiagonana__gte=nanoboka, TaonaNahatogavanaTaoAmpiagonana__lt=nifarana)
    if Mpadray:
        mpino = mpino.filter(M_P="P")

    # Pagination
    paginator = Paginator(mpino, par_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    # Convertir les données en JSON
    data = {
        "Mpino": [
            {
                "Anarana": mpino.Anarana,
                "Fanampiny" : mpino.Fanampiny,
                "mpino" : mpino,
                "Teraka": mpino.Teraka,

                "TaonaNahatogavanaTaoAmpiagonana": mpino.TaonaNahatogavanaTaoAmpiagonana,

            }
            for mpino in page_obj
        ],
        "pagination": {
            "current_page": page_obj.number,
            "total_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
            "has_previous": page_obj.has_previous(),
        }
    }

    context = {"moba": moba, "fikaroana": fikaroana, "ita": ita, "isa": isa,'data': data

               }

    return render(request, 'tableau/Mpino.html', context)

def import_Mpino_Exel(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active  # Accéder à la première feuille

            # Vérifier l'entête et ajouter les lignes

            for row in sheet.iter_rows(min_row=2, values_only=True):
                #LaharanaKaratra, Anarana, Fanampiny, Teraka, L_v, Antonasa, Adresy, Faritra, Fiday, TaonaNahatogavanaTaoAmpiagonana, FiagonanaNiavina, batisa, taona_nanaovana_batisa,M_P, TaonaNahaMpandray, DatyNakanaKaratra = row
                #N,MEN,CF, Anarana , Fanampiny, Lot,Faritra, TNF,TN,TP, L_V,Batisa, TNB,  Mpandray,asa,  tel,NAKANA_KARATRA, fg_niaviana= row
                N, MEN, CF, Anarana, Fanampiny, Lot, Faritra, TNF, TN, TP, L_V, Batisa, Mpandray, asa, tel,fg_niaviana, NAKANA_KARATRA  = row
                Anarana = Anarana or " "
                Fanampiny = Fanampiny or " "
                Lot = Lot or " "
                Faritra = Faritra or " "
                try:
                    TNF = int(TNF)
                except:
                    TNF = 0

                TN = TN or 0
                try:
                    TP = int(TP)
                except:
                    TP = 0
                if str(L_V).lower()=="lahy":
                    L_V_ = "L"
                elif str(L_V).lower()=="vavy":
                    L_V_ = "V"

                if str(Batisa).lower()== "eny":
                    Batisa = "ENY"
                elif str(Batisa).lower()== "tsy" or str(Batisa).lower()== "tsia":
                    Batisa = "TSY"
                else: Batisa = "TSY"
                if str(Mpandray).lower() == "eny":
                    Mpandray_= "P"
                elif str(Mpandray).lower() == "tsy" or str(Mpandray).lower() == "tsia":
                    Mpandray_ = "K"
                asa = asa or " "
                tel = str(tel) or " "
                fg_niaviana = fg_niaviana or " "
                NAKANA_KARATRA = NAKANA_KARATRA or " "

                if not Mpino3.objects.filter(Anarana=Anarana, Fanampiny=Fanampiny,LaharanaKaratra="/".join([str(N),str(MEN),str(CF)])).exists():
                    Mpino3.objects.create(
                        LaharanaKaratra="/".join([str(N),str(MEN),str(CF)]),
                        Anarana=Anarana,
                        Fanampiny=Fanampiny,
                        Teraka=int(TN),


                        L_v=L_V_,
                        Antonasa=asa,

                        Adresy=Lot,
                        Faritra=Faritra,
                        Fiday=tel,

                        TaonaNahatogavanaTaoAmpiagonana=int(TNF),
                        FiagonanaNiavina=fg_niaviana,
                        batisa=Batisa,
                        taona_nanaovana_batysa= 0,
                        M_P=Mpandray_,
                        TaonaNahaMpandray=int(TP),

                        DatyNakanaKaratra=NAKANA_KARATRA,

                    )
                else:
                    LaharanaKaratra = "/".join([str(N),str(MEN),str(CF)])
                    messages.success(request, f"{Anarana},{Fanampiny},{LaharanaKaratra} Efa misy")


            messages.success(request, "Les données ont été importées avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect('Tableau:import_Mpino_Exel')
    return render(request, 'tableau/import_excel.html')


def export_Mpino_Exel(request):
    # Récupérer les données
    Mpino = Mpino3.objects.all()

    # Récupérer tous les paramètres de filtre et de recherche
    fikaroana = request.GET.get('q', '')
    taona_nahatogavana = request.GET.get('taona', '')
    nanoboka = request.GET.get('nanoboka', '')
    nifarana = request.GET.get('nifarana', '')

    Mpadray = request.GET.get('mpadray', '')
    # Filtre
    if taona_nahatogavana:
        Mpino= Mpino.filter(TaonaNahatogavanaTaoAmpiagonana=taona_nahatogavana)
    if nanoboka and nifarana:
        Mpino = Mpino.filter(TaonaNahatogavanaTaoAmpiagonana__gte=nanoboka,
                             TaonaNahatogavanaTaoAmpiagonana__lt=nifarana)
    if Mpadray:
        Mpino = Mpino.filter(M_P="P")
    # Créer un fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MpinoNavokag"

    # Ajouter l'entête

    #header = ['LaharanaKaratra', 'Anarana', 'Fanampiny', 'Teraka', 'L_v', 'Antonasa',  'Adresy', 'Faritra', 'Fiday', 'TaonaNahatogavanaTaoAmpiagonana', 'FiagonanaNiavina', 'batisa',' taona_nanaovana_batisa','M_P',
    #          'TaonaNahaMpandray', 'DatyNakanaKaratra']
    header = ['N','MEN','CF', 'Anarana', "Fanampiny",'Lot', 'Faritra','TNF','TN','TP','L/V','batisa', ' TNB','Mpandray', 'asa',  'tel'
              ,'NAKANA_KARATRA', 'fg_niaviana'
              ]
    ws.append(header)
   # for cell in ws[1]:
    #    cell.font = Font(bold=True)

    # Ajouter les données
    for mpino in Mpino:
        Lk = str(mpino.LaharanaKaratra).split("/")
        ws.append(
            [Lk[0],Lk[1],Lk[2], mpino.Anarana, mpino.Fanampiny,  mpino.Adresy,mpino.Faritra,  mpino.TaonaNahatogavanaTaoAmpiagonana,
             mpino.Teraka, mpino.TaonaNahaMpandray,mpino.L_v, mpino.batisa,
             mpino.taona_nanaovana_batysa,mpino.M_P,mpino.Antonasa, mpino.Fiday,

              mpino.DatyNakanaKaratra, mpino.FiagonanaNiavina,])

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=mpinoNavoakag.xlsx'

    # Sauvegarder le fichier dans la réponse
    wb.save(response)
    return response

def import_Adidy_Exel(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active  # Accéder à la première feuille

            # Vérifier l'entête et ajouter les lignes

            for row in sheet.iter_rows(min_row=2, values_only=True):
                Anarana,Fanampiny,LaharanyKaratra,Mpadray,Mpino, Taona, J, P_J, J_Date, F, P_F, F_Date, M, P_M, M_Date, Av, P_Av, Av_Date, Ma, P_Ma, Ma_Date, Ji, P_Ji, Ji_Date, Ju, P_Ju, Ju_Date, Ag, P_Ag, Ag_Date, S, P_S, S_Date, Ak, P_Ak, Ac_Date,N, P_N, N_Date, D, P_D, D_Date = row
                Adidy.objects.create(
                Mpino = Mpino3(id=Mpino),
                Taona = Taona,

                J = J,
                P_J = P_J,
                J_Date = J_Date,

                F = F,
                P_F =P_F ,
                F_Date = F_Date,

                M = M,
                P_M = P_M,
                M_Date =M_Date ,

                Av =Av,
                P_Av =P_Av,
                Av_Date = Av_Date,

                Ma = Ma,
                P_Ma = P_Ma,
                Ma_Date =Ma_Date,

                Ji = Ji,
                P_Ji = P_Ji,
                Ji_Date =Ji_Date,

                Ju = Ju,
                P_Ju =  P_Ju,
                Ju_Date = Ju_Date,

                Ag = Ag,
                P_Ag = P_Ag,
                Ag_Date = Ag_Date,

                S = S ,
                P_S = P_S,
                S_Date = S_Date,

                Ak =Ak ,
                P_Ak = P_Ak,
                Ac_Date =Ac_Date ,

                N =  N,
                P_N = P_N,
                N_Date =N_Date ,

                D = D ,
                P_D =P_D ,
                D_Date =  D_Date,

                )

            messages.success(request, "Les données ont été importées avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect('Tableau:import_Adidy_Exel')
    return render(request, 'tableau/import_excel.html')

def export_Adidy_Exell(request):
    # Récupérer les données
    adidy = Adidy.objects.all()

    # Créer un fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AdidyMivoka"

    # Ajouter l'entête
    header = ['Anarana','Fanampiny','LaharanyKaratra','Mpadray(P)/Kristianiana(k)','Mpino','Taona','J','P_J','J_Date','F','P_F','F_Date','M','P_M','M_Date','Av','P_Av','Av_Date',
                  'Ma','P_Ma','Ma_Date','Ji','P_Ji','Ji_Date','Ju','P_Ju','Ju_Date','Ag','P_Ag','Ag_Date','S','P_S','S_Date','Ak','P_Ak','Ac_Date',
                  'N','P_N','N_Date','D','P_D','D_Date']
    ws.append(header)
   # for cell in ws[1]:
    #    cell.font = Font(bold=True)

    # Ajouter les données
    for addy in adidy:
        ws.append(
            [addy.Mpino.Anarana,addy.Mpino.Fanampiny,addy.Mpino.LaharanaKaratra,addy.Mpino.M_P,int(addy.Mpino.id), addy.Taona, addy.J, addy.P_J, addy.J_Date, addy.F, addy.P_F, addy.F_Date, addy.M, addy.P_M, addy.M_Date, addy.Av, addy.P_Av,
             addy.Av_Date,
             addy.Ma, addy.P_Ma, addy.Ma_Date, addy.Ji, addy.P_Ji, addy.Ji_Date, addy.Ju, addy.P_Ju, addy.Ju_Date, addy.Ag, addy.P_Ag, addy.Ag_Date, addy.S,
             addy.P_S, addy.S_Date, addy.Ak, addy.P_Ak, addy.Ac_Date,
             addy.N, addy.P_N, addy.N_Date, addy.D, addy.P_D, addy.D_Date])

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=AdidyMivoka.xlsx'

    # Sauvegarder le fichier dans la réponse
    wb.save(response)
    return response

def Adidy_t(request):
    #adidy = Adidy.objects.all().order_by('Taona')

    # Récupérer tous les paramètres de filtre et de recherche
    fikaroana = request.GET.get('q', '')
    taona = request.GET.get('taona', '')
    #nanoboka = request.GET.get('nanoboka', '')
    #nifarana = request.GET.get('nifarana', '')
    par_page = int(request.GET.get('par_page', 10))

    # Filtrer les cotisations
    adidy = Adidy.objects.select_related('Mpino').all().order_by('Taona')
    if taona :
        adidy = adidy.filter(Taona=taona )
    if fikaroana:
        adidy = adidy.filter(Mpino__Anarana__icontains=fikaroana) or adidy.filter(Mpino__LaharanaKaratra__icontains=str(fikaroana))

    # Pagination
    paginator = Paginator(adidy, par_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Préparer les données pour JSON
    data = {
        "Adidy": [
            {   "adidy" : adidy,

            }
            for adidy in page_obj
        ],
        "pagination": {
            "current_page": page_obj.number,
            "total_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
            "has_previous": page_obj.has_previous(),
        },
    }


    context = {"adidy":adidy,'data': data}


    return render(request,'tableau/Adidy.html',context)

def Katekomen_L(request):
    #adidy = Adidy.objects.all().order_by('Taona')

    # Récupérer tous les paramètres de filtre et de recherche
    fikaroana = request.GET.get('q', '')
    taona = request.GET.get('taona', '')
    #nanoboka = request.GET.get('nanoboka', '')
    #nifarana = request.GET.get('nifarana', '')
    par_page = int(request.GET.get('par_page', 10))

    # Filtrer les cotisations
    katekomen = Katekomen.objects.select_related('Mpino').all().order_by('Taompianarana')
    if taona :
        katekomen = katekomen.filter(Taompianarana=taona )
    if fikaroana:
        katekomen = katekomen.filter(Mpino__Anarana__icontains=fikaroana) or adidy.filter(Mpino__LaharanaKaratra__icontains=str(fikaroana))

    # Pagination
    paginator = Paginator(katekomen, par_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Préparer les données pour JSON
    data = {
        "isa": katekomen.count(),
        "Katekomen": [
            {   "katekomen" : katekomen,


            }
            for katekomen in page_obj
        ],
        "pagination": {
            "current_page": page_obj.number,
            "total_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
            "has_previous": page_obj.has_previous(),
        },
    }


    context = {"katekomen":katekomen,'data': data}


    return render(request,'tableau/Katekomen.html',context)


def TobanaMpino(request):
    moba = Mpino3.objects.all()
    isa = moba.count()
    Lehilay = Mpino3.objects.filter(L_v="L").count()
    Vehivavy = Mpino3.objects.filter(L_v="V").count()
    salanisaL = round((Lehilay * 100) / isa)
    salanisaV = round((Vehivavy * 100) / isa)

    padray = Mpino3.objects.filter(M_P="P").count()
    kristianina = Mpino3.objects.filter(M_P="K").count()

    # Récupération des filtres depuis le formulaire
    annee = request.GET.get('annee')  # Filtrer par année donnée
    debut = request.GET.get('debut')  # Début période
    fin = request.GET.get('fin')  # Fin période

    # Filtrage des personnes
    mpino = Mpino3.objects.all()
    if annee:
        mpino = mpino.filter(TaonaNahatogavanaTaoAmpiagonana=annee)
    elif debut and fin:
        mpino = mpino.filter(TaonaNahatogavanaTaoAmpiagonana__range=(debut, fin))

    # Calcul des statistiques
    total_inscrits = mpino.count()
    hommes = mpino.filter(L_v='L').count()
    femmes = mpino.filter(L_v='V').count()
    pourcentage_hommes = (hommes / total_inscrits * 100) if total_inscrits > 0 else 0
    pourcentage_femmes = (femmes / total_inscrits * 100) if total_inscrits > 0 else 0

    mpadray =  mpino.filter(M_P='P').count()
    kristianina =  mpino.filter(M_P='K').count()
    pourcentage_mpadray = (mpadray / total_inscrits * 100) if total_inscrits > 0 else 0
    pourcentage_kristianina = (kristianina/ total_inscrits * 100) if total_inscrits > 0 else 0

    vita_batisa = mpino.filter(batisa="ENY").count()
    tsy_vita_batisa = mpino.filter(batisa="TSY").count()
    pourcentage_vita_batisa = (vita_batisa / total_inscrits * 100) if total_inscrits > 0 else 0
    pourcentage_tsy_vita_batisa = (tsy_vita_batisa / total_inscrits * 100) if total_inscrits > 0 else 0

    tranche_age = {
        'entre_1_10': mpino.filter(
            Teraka__lt=(date.today().year),
            Teraka__gte=(date.today().year )
        ).count(),
        'entre_10_15': mpino.filter(
            Teraka__lt=(date.today().year - 10),
            Teraka__gte=(date.today().year - 15)
        ).count(),

        'entre_15_18': mpino.filter(
            Teraka__lt=(date.today().year - 15),
            Teraka__gte=(date.today().year - 18)
        ).count(),

        'moins_18': mpino.filter(Teraka__gte=(date.today().year - 18)).count(),

        'entre_18_25': mpino.filter(
            Teraka__lt=(date.today().year - 18),
            Teraka__gte=(date.today().year - 25)
        ).count(),

        'plus_25': mpino.filter(Teraka__lt=(date.today().year - 25)).count(),

        'entre_25_30': mpino.filter(
            Teraka__lt=(date.today().year - 25),
            Teraka__gte=(date.today().year - 30)
        ).count(),

        'entre_30_40': mpino.filter(
            Teraka__lt=(date.today().year - 30),
            Teraka__gte=(date.today().year - 40)
        ).count(),

        'entre_40_50': mpino.filter(
            Teraka__lt=(date.today().year - 40),
            Teraka__gte=(date.today().year - 50)
        ).count(),

        'plus_50': mpino.filter(Teraka__lt=(date.today().year - 50)).count(),
    }

    # Préparation des données JSON pour les statistiques
    statistiques_data = {
        'total_inscrits': total_inscrits,
        'Lehilay':hommes,
        'Vehivavy':femmes,
        'pourcentage_hommes': round(pourcentage_hommes, 2),
        'pourcentage_femmes': round(pourcentage_femmes, 2),
        'tranche_age': tranche_age,

        'mpadray' : mpadray,
        'kristianina' : kristianina,
        'pourcentage_mpadray' : pourcentage_mpadray,
        'pourcentage_kristianina' : pourcentage_kristianina,

        'vita_batisa' :vita_batisa,
        'tsy_vita_batisa' : tsy_vita_batisa,
        'pourcentage_vita_batisa' : pourcentage_vita_batisa,
        'pourcentage_tsy_vita_batisa' : pourcentage_tsy_vita_batisa,

    }

    # Récupérer d'autres informations à passer au template
   # autres_infos = AutreModele.objects.all()

    # Retourner le template avec les données
    context = {
        #'autres_infos': autres_infos,
        'statistiques_json': statistiques_data,
        'annee': annee,
        'debut': debut,
        'fin': fin,

        "isa": isa, "Lehilay": hommes, "Vehivavy": femmes,
         "padray": padray, "kristianina": kristianina
    }
    return render(request, 'tableau/state.html', context)

def TobanaKatekomen(request):
    moba = Katekomen.objects.all()
    isa = moba.count()
    Lehilay = Katekomen.objects.filter(Mpino__L_v="L").count()
    Vehivavy = Katekomen.objects.filter(Mpino__L_v="V").count()
    salanisaL = round((Lehilay * 100) / isa)
    salanisaV = round((Vehivavy * 100) / isa)

    #padray = Mpino3.objects.filter(M_P="P").count()
    #kristianina = Mpino3.objects.filter(M_P="K").count()

    # Récupération des filtres depuis le formulaire
    annee = request.GET.get('annee')  # Filtrer par année donnée
    debut = request.GET.get('debut')  # Début période
    fin = request.GET.get('fin')  # Fin période

    # Filtrage des personnes
    katekomen = Katekomen.objects.all()
    if annee:
        katekomen = katekomen.filter(Taompianarana=annee)
    elif debut and fin:
        katekomen = katekomen.filter(Daty_Nidirana=(debut, fin))

    # Calcul des statistiques
    total_inscrits = katekomen.count()
    hommes = Katekomen.objects.filter(Mpino__L_v='L').count()
    femmes = Katekomen.objects.filter(Mpino__L_v='V').count()
    pourcentage_hommes = (hommes / total_inscrits * 100) if total_inscrits > 0 else 0
    pourcentage_femmes = (femmes / total_inscrits * 100) if total_inscrits > 0 else 0

    #mpadray =  mpino.filter(M_P='P').count()
    #kristianina =  mpino.filter(M_P='K').count()
    #pourcentage_mpadray = (mpadray / total_inscrits * 100) if total_inscrits > 0 else 0
    #pourcentage_kristianina = (kristianina/ total_inscrits * 100) if total_inscrits > 0 else 0

    #vita_batisa = mpino.filter(batisa="ENY").count()
    #tsy_vita_batisa = mpino.filter(batisa="TSY").count()
    #pourcentage_vita_batisa = (vita_batisa / total_inscrits * 100) if total_inscrits > 0 else 0
    #pourcentage_tsy_vita_batisa = (tsy_vita_batisa / total_inscrits * 100) if total_inscrits > 0 else 0

    tranche_age = {
        'entre_1_10': katekomen.filter(
            Mpino__Teraka__lt=(date.today().year),
            Mpino__Teraka__gte=(date.today().year )
        ).count(),
        'entre_10_15': katekomen.filter(
            Mpino__Teraka__lt=(date.today().year - 10),
            Mpino__Teraka__gte=(date.today().year - 15)
        ).count(),

        'entre_15_18': katekomen.filter(
            Mpino__Teraka__lt=(date.today().year - 15),
            Mpino__Teraka__gte=(date.today().year - 18)
        ).count(),

        'moins_18': katekomen.filter(Mpino__Teraka__gte=(date.today().year - 18)).count(),

        'entre_18_25': katekomen.filter(
            Mpino__Teraka__lt=(date.today().year - 18),
            Mpino__Teraka__gte=(date.today().year - 25)
        ).count(),

        'plus_25': katekomen.filter(Mpino__Teraka__lt=(date.today().year - 25)).count(),

        'entre_25_30': katekomen.filter(
            Mpino__Teraka__lt=(date.today().year - 25),
            Mpino__Teraka__gte=(date.today().year - 30)
        ).count(),

        'entre_30_40': katekomen.filter(
            Mpino__Teraka__lt=(date.today().year - 30),
            Mpino__Teraka__gte=(date.today().year - 40)
        ).count(),

        'entre_40_50': katekomen.filter(
            Mpino__Teraka__lt=(date.today().year - 40),
            Mpino__Teraka__gte=(date.today().year - 50)
        ).count(),

        'plus_50': katekomen.filter(Mpino__Teraka__lt=(date.today().year - 50)).count(),
    }

    # Préparation des données JSON pour les statistiques
    statistiques_data = {
        'total_inscrits': total_inscrits,
        'pourcentage_hommes': round(pourcentage_hommes, 2),
        'pourcentage_femmes': round(pourcentage_femmes, 2),
        'tranche_age': tranche_age,

    }

    # Retourner le template avec les données
    context = {
        #'autres_infos': autres_infos,
        'statistiques_json': statistiques_data,
        'annee': annee,
        'debut': debut,
        'fin': fin,

        "isa": isa, "Lehilay": hommes, "Vehivavy": femmes,

    }
    return render(request, 'tableau/TobanaKatekomen.html', context)
def Tobana_Arabola(request):
    # Récupérer les paramètres de filtre
    annee = request.GET.get('annee')
    debut = request.GET.get('debut')
    fin = request.GET.get('fin')
    comparaison_annee = request.GET.get('comparaison_annee')


    # Filtrage des cotisations
    adidy = Adidy.objects.select_related('Mpino').order_by('Taona')
    #adidy = adidy.filter(Taona=taona)
    if annee:
        taona = annee
        adidy = adidy.filter(Taona=annee)
    else:
        taona = date.today().year
        adidy = adidy.filter(Taona=taona)
    """
    elif debut and fin:
        adidy = adidy.filter(Taona__range=[debut, fin])
    """
    # Calcul des statistiques pour la période sélectionnée
    total_personnes = adidy.values('Mpino').distinct().count()

    total_cotisations = adidy.aggregate(
        total=Sum('Fitabarany')
    )['total'] or 0
    """
    #total_cotisations = adidy.values('Fitabarany').distinct().count()
    total_cotisations_theorique = total_personnes * 12
    pourcentage_paiement = (
        (total_cotisations / total_cotisations_theorique) if total_cotisations_theorique > 0 else 0
    )
    """
    personnes_tout_paye = adidy.filter(
        Q(J__gt=0) & Q(F__gt=0) & Q(M__gt=0) &
        Q(Av__gt=0) & Q(Ma__gt=0) & Q(Ji__gt=0) &
        Q(Ju__gt=0) & Q(Ag__gt=0) & Q(S__gt=0) &
        Q(Ak__gt=0) & Q(N__gt=0) & Q(D__gt=0)
    ).values('Mpino').distinct().count()
    """
    moyenne_cotisations = adidy.aggregate(
        avg=Avg('Fitabarany'))['avg'] or 0
    """
    # Comparaison avec une autre année
    total_cotisations_comparaison = 0
    moyenne_cotisations_comparaison = 0
    personnes_tout_paye_comparaison = 0
    pourcentage_paiement_comparaison = 0

    if comparaison_annee:
        adidy = Adidy.objects.filter(Taona=annee)
        comparaison_cotisations = Adidy.objects.filter(Taona=comparaison_annee)
        total_cotisations_comparaison = comparaison_cotisations.aggregate(
            total=Sum('Fitabarany'))['total'] or 0
        """
        total_comparaison_personnes = comparaison_cotisations.values('Mpino').distinct().count()
        total_comparaison_cotisations_theorique = total_comparaison_personnes * 12
        pourcentage_paiement_comparaison = (
            (total_cotisations_comparaison / total_comparaison_cotisations_theorique) * 100
            if total_comparaison_cotisations_theorique > 0 else 0
        )

        moyenne_cotisations_comparaison = comparaison_cotisations.aggregate(
            avg=Avg('Fitabarany'))['avg'] or 0
        """
        personnes_tout_paye_comparaison = comparaison_cotisations.filter(
            Q(J__gt=0) & Q(F__gt=0) & Q(M__gt=0) &
            Q(Av__gt=0) & Q(Ma__gt=0) & Q(Ji__gt=0) &
            Q(Ju__gt=0) & Q(Ag__gt=0) & Q(S__gt=0) &
            Q(Ak__gt=0) & Q(N__gt=0) & Q(D__gt=0)
        ).values('Mpino').distinct().count()

    #-----------------------------------------------------

    #par_page = int(request.GET.get('par_page', 5))
    resetes = VolaNiditra.objects.all().order_by('-Taona')
    depenses = VolaNivoka.objects.all().order_by('-Taona')
    if annee:
        resetes = resetes.filter(Taona=annee)
        depenses = depenses.filter(Taona=annee)

    # Somme des resetes, cotisations, et dépenses pour l'année spécifiée
    somme_resetes = resetes.aggregate(total=Sum('Vola'))['total'] or 0
    somme_depenses = depenses.aggregate(total=Sum('Vola'))['total'] or 0


    Somme = (int(total_cotisations)+ int( somme_resetes)) - int(somme_depenses)
    """
    # Pagination pour resetes et dépenses
    from django.core.paginator import Paginator
    resetes_paginator = Paginator(resetes, par_page)
    depenses_paginator = Paginator(depenses, par_page)
    resetes_page = resetes_paginator.get_page(request.GET.get('page_resetes', 5))
    depenses_page = depenses_paginator.get_page(request.GET.get('page_depenses', 5))


    # Pagination resetes
    paginator = Paginator(resetes, par_page)
    page_number = request.GET.get('page',5)
    page_obj = paginator.get_page(page_number)

    # Pagination depences
    paginator_r = Paginator(depenses, par_page)
    page_number_r = request.GET.get('page_r',5)
    page_obj_r = paginator.get_page(page_number)
    """
    # Préparer les données pour le rendu
    data = {
        "resetes": resetes,
        "depenses": depenses,
        "Somme":Somme,

        "totaux": {
            "somme_resetes": somme_resetes,

            "somme_depenses": somme_depenses,

        },

       # "pagination": {
        #    "current_page": page_obj.number,
        #    "total_pages": paginator.num_pages,
        #    "has_next": page_obj.has_next(),
        #    "has_previous": page_obj.has_previous(),

        #    "current_page_r": page_obj_r.number,
       #     "total_pages_r": paginator_r.num_pages,
        #    "has_next_r": page_obj_r.has_next(),
        #    "has_previous_r": page_obj_r.has_previous(),
       # }
    }

    #------------------------------------------------------

    # Préparation des statistiques sous forme de JSON
    statistiques = {
        'taona': taona,
        "periode_selectionnee": {

            "somme_cotisations": total_cotisations,
           # "moyenne_cotisations": moyenne_cotisations,
            "personnes_tout_paye": personnes_tout_paye,
            #"pourcentage_paiement": pourcentage_paiement,
        },
        "comparaison": {
            "annee": comparaison_annee,
            "somme_cotisations": total_cotisations_comparaison,
           # "moyenne_cotisations": moyenne_cotisations_comparaison,
            "personnes_tout_paye": personnes_tout_paye_comparaison,
           # "pourcentage_paiement": pourcentage_paiement_comparaison,
        }
    }

    context = {
        "data":data,
        #"statistiques_json": json.dumps(statistiques),
        "statistiques": statistiques,
        "comparaison_annee": comparaison_annee,
        "annee": annee,
        "debut": debut,
        "fin": fin,
    }

    return render(request, 'tableau/Tobana_Arabola.html', context)

def export_Volaniditra(request,taona):
    # Récupérer les données
    Niditra = VolaNiditra.objects.filter(Taona=taona)

    # Créer un fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VolaNiditra"

    # Ajouter l'entête
    header = ['Taona', 'Daty', 'Antony', 'Vola', 'Fanamarihana']
    ws.append(header)
   # for cell in ws[1]:
    #    cell.font = Font(bold=True)

    # Ajouter les données
    for niditra in Niditra:
        ws.append(
            [niditra.Taona,niditra.Daty,niditra.Antony,niditra.Vola,niditra.Fanamarihana])

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; VolaNiditra.xlsx'

    # Sauvegarder le fichier dans la réponse
    wb.save(response)
    return response

def export_Volanivoka(request,taona):
    # Récupérer les données
    Nivoka = VolaNivoka.objects.filter(Taona=taona)

    # Créer un fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VolaNivoka"

    # Ajouter l'entête
    header = ['Taona', 'Daty', 'Antony', 'Vola', 'Fanamarihana']
    ws.append(header)
   # for cell in ws[1]:
    #    cell.font = Font(bold=True)

    # Ajouter les données
    for nivoka in Nivoka:
        ws.append(
            [nivoka.Taona,nivoka.Daty,nivoka.Antony,nivoka.Vola,nivoka.Fanamarihana])

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; VolaNiditra.xlsx'

    # Sauvegarder le fichier dans la réponse
    wb.save(response)
    return response

def Ijery(request, anarana):
    Anarana  = get_object_or_404(Mpino3,pk=anarana)

    fanamarihana = Anarana.F_Mpino.all() # F_Mpino -> related_name dans le models.py
    if request.method == "POST":
        form = FanamarihanaForm(request.POST)
        if form.is_valid():
            fanamarihana_ = form.save(commit=False)
            fanamarihana_.F_Mpino = Anarana
            fanamarihana_.save()
            return redirect("Tableau:Ijery",anarana=anarana)
    else:
        form = FanamarihanaForm()
    context = {"anarana": Anarana, "fanamarihana": fanamarihana,"form":form}
    return render(request, 'tableau/Ijery.html', context)

def Ampiditra_Fanamarihana(request,anarana):
    if request.method == "POST":
        titre = request.POST.get("titre")
        contenu = request.POST.get("contenu")
        if titre and contenu:
            article = Fanamarihana.objects.create(F_Mpino=anarana, Fanamarihana=contenu)
            return JsonResponse({
                "id": article.id,
                "titre": article.titre,
                "contenu": article.contenu
            })
    return JsonResponse({"error": "Données invalides"}, status=400)


def Amafa_Fanamarihana(request, anarana):
    fanamarihana = get_object_or_404(Fanamarihana, pk=anarana)
    personne_id = fanamarihana.F_Mpino.id # On récupère l'ID de la personne avant suppression
    fanamarihana.delete()
    return redirect("Tableau:Ijery",anarana=personne_id)


